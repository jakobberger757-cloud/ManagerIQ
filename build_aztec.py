"""
Aztec Group — US Prospecting Intelligence Tool
Standalone Build Script

USAGE:
    python3 build_aztec.py

REQUIREMENTS:
    pip install pandas openpyxl

INPUT FILES (place in same folder as this script, or update paths below):
    convergence_managers.csv      — Convergence manager-level export
    convergence_funds.csv         — Convergence fund-level export
    preqin_managers.xlsx          — Preqin fund manager export (US + Canada)
    preqin_funds.xlsx             — Preqin fund-level export
    preqin_forecast.xlsx          — Preqin fund forecast export

OUTPUT:
    aztec_prospecting.html        — Self-contained tool, open in Chrome/Safari

UPDATING DATA:
    1. Export fresh files from Convergence and Preqin
    2. Rename them to match the filenames above
    3. Run: python3 build_aztec.py
    4. Open the new aztec_prospecting.html

UPDATING CLIENT/TARGET LISTS:
    Edit the AZTEC_CLIENTS and CRM_TARGETS lists in the CONFIG section below.

VERSION: Built from conversation session March 2026
"""

import os, json, re, warnings
import pandas as pd
from collections import Counter
from openpyxl import load_workbook
warnings.filterwarnings('ignore')

# ══════════════════════════════════════════════════════════════════
# CONFIG — edit these as your lists change
# ══════════════════════════════════════════════════════════════════

# Paths to input files (relative to this script, or use absolute paths)
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))

FILES = {
    'conv_managers': os.path.join(SCRIPT_DIR, 'convergence_managers.csv'),
    'conv_funds':    os.path.join(SCRIPT_DIR, 'convergence_funds.csv'),
    'pq_managers':   os.path.join(SCRIPT_DIR, 'preqin_managers.xlsx'),
    'pq_funds':      os.path.join(SCRIPT_DIR, 'preqin_funds.xlsx'),
    'pq_forecast':   os.path.join(SCRIPT_DIR, 'preqin_forecast.xlsx'),
}

OUTPUT_FILE = os.path.join(SCRIPT_DIR, 'aztec_prospecting.html')

TODAY = pd.Timestamp('2026-03-23')  # Update to today's date when refreshing

# Aztec global clients (305 firms as of March 2026)
# Add new wins, remove churned clients
AZTEC_CLIENTS = [
    "Clearlake Capital Group","H.I.G. Capital","Nuveen","Advent International",
    "LaSalle Investment Management","StepStone Group","Resource Capital Funds",
    "Principal Asset Management","Warburg Pincus","Apax Partners","SoftBank",
    "Schroders Capital","Hollyport Capital","FHR Capital","Water Equity",
    "Harbert Management","Leeds Equity Partners","Lovell Minnick Partners",
    "CCMP Capital","Oaktree Capital Management","Thoma Bravo","Adams Street Partners",
    "Bain Capital","KKR","Fortress Investment Group","TA Associates","PGIM",
    "Blackstone Strategic Partners","ArcLight Capital Partners","Corebridge Financial",
    "Energy Capital Partners","GenNx360","Macquarie Asset Management","PSP Investments",
    "OpenGate Capital",  # NOTE: depositary relationship only, not full admin client
    # Add more clients here as you win them
]

# US-confirmed clients (subset of above)
AZTEC_US_CLIENTS = [
    "Clearlake Capital Group","H.I.G. Capital","Nuveen","StepStone Group",
    "Resource Capital Funds","FHR Capital","Water Equity","Harbert Management",
    "Leeds Equity Partners","Lovell Minnick Partners","CCMP Capital",
    "Oaktree Capital Management","Thoma Bravo","Adams Street Partners",
    "ArcLight Capital Partners","Energy Capital Partners","GenNx360",
    "PSP Investments","Warburg Pincus",
]

# OpenGate = existing UK depositary relationship, not full admin client
OPENGATE_NAME = "OpenGate Capital"

# CRM targets (Jakob's active pipeline — 91 firms as of March 2026)
CRM_TARGETS = [
    "Pearl Energy Investments","MiddleGround Capital","Mill Point Capital",
    "Align Capital Partners","The Halifax Group","Summit Park",
    "Avista Healthcare","Capstreet","Carousel Capital","Castle Creek Capital",
    "Chicago Pacific Founders","Edgewater Funds","Windjammer Capital",
    "Irving Place Capital","Ocean Avenue Capital","Silver Oak Equity",
    "Huron Capital Partners","RRA Advisors","Titanium Equity Partners",
    "Vance Street Capital","TorQuest Partners","Birch Hill Equity Partners",
    "Clairvest GP Manageco","Vectors Capital","Aqualis Partners",
    # Add your full 91-firm list here
]

# CRM notes for specific firms
CRM_NOTES = {
    "Pearl Energy Investments": "CFO meeting scheduled March 24 2026",
    "MiddleGround Capital": "Briefly spoken to CFO on cold call",
    "Mill Point Capital": "Head of IR: not changing providers in immediate future",
    "Align Capital Partners": "CFO May 2025: all good for now. Continue to nurture",
    "The Halifax Group": "Recently transitioned to outsourced admin. Longer play",
    "Halifax": "Recently transitioned to outsourced admin. Longer play",
    "Edgewaters": "Recently transitioned to outsourced admin. Longer play",
    "Summit Park": "Meeting scheduled June 2026",
    "TorQuest Partners": "Canadian PE — cross-border expansion opportunity",
    "Birch Hill Equity Partners": "Canadian PE — cross-border expansion opportunity",
    "Clairvest GP Manageco": "Canadian PE — cross-border expansion opportunity",
}

# ══════════════════════════════════════════════════════════════════
# NORMALISATION HELPERS
# ══════════════════════════════════════════════════════════════════

def norm(s):
    """Normalise firm name for matching across sources."""
    if not s: return ''
    s = str(s).upper().strip()
    for w in [' LLC',' LP',' INC',' GROUP',' PARTNERS',' CAPITAL',
              ' MANAGEMENT',' ADVISORS',' ADVISOR',' FUND',' SERVICES',
              ' REAL ESTATE',' INVESTMENTS',' INVESTMENT',' & CO',
              ' &',' CO.',' CORP',' THE',' ASSOCIATES',' EQUITY',
              ' VENTURES',' NORTH AMERICA',' MANAGECO',' ADVISERS',
              ' MANAGERS',' HOLDINGS',' ASSET',' CORPORATION']:
        s = s.replace(w, '')
    return re.sub(r'[^A-Z0-9 ]', '', s).strip()

def clean_name(s):
    """Strip entity suffixes from display names."""
    if not s: return ''
    s = str(s).strip()
    for suffix in [', LLC',', L.L.C.',', L.L.C',' LLC',', LP',', L.P.',
                   ' LP',', INC',', INC.',' INC',', LTD',', LTD.',' LTD',
                   ', CORP',' CORP',', L.L.P.',', LLP']:
        if s.upper().endswith(suffix.upper()):
            s = s[:-len(suffix)].rstrip(',').strip()
    return s

def tc(s):
    """Title-case a cleaned name."""
    if not s: return ''
    return ' '.join(w.capitalize() for w in clean_name(str(s)).strip().split())

def clean(v):
    """Return None for blank/null values."""
    if v is None or (isinstance(v, float) and pd.isna(v)): return None
    s = str(v).strip()
    return None if s in ['-','nan','NaN','','None','N/A','0'] else s

def norm_state(s):
    if not s: return None
    s = str(s).strip()
    if s == 'Canada': return 'Canada'
    STATE_MAP = {
        'ALABAMA':'AL','ALASKA':'AK','ARIZONA':'AZ','ARKANSAS':'AR','CALIFORNIA':'CA',
        'COLORADO':'CO','CONNECTICUT':'CT','DELAWARE':'DE','DELWARE':'DE','FLORIDA':'FL',
        'GEORGIA':'GA','HAWAII':'HI','IDAHO':'ID','ILLINOIS':'IL','INDIANA':'IN',
        'IOWA':'IA','KANSAS':'KS','KENTUCKY':'KY','LOUISIANA':'LA','MAINE':'ME',
        'MARYLAND':'MD','MASSACHUSETTS':'MA','MICHIGAN':'MI','MINNESOTA':'MN',
        'MISSISSIPPI':'MS','MISSOURI':'MO','MONTANA':'MT','NEBRASKA':'NE','NEVADA':'NV',
        'NEW HAMPSHIRE':'NH','NEW JERSEY':'NJ','NEW MEXICO':'NM','NEW YORK':'NY',
        'NORTH CAROLINA':'NC','NORTH DAKOTA':'ND','OHIO':'OH','OKLAHOMA':'OK',
        'OREGON':'OR','PENNSYLVANIA':'PA','RHODE ISLAND':'RI','SOUTH CAROLINA':'SC',
        'SOUTH DAKOTA':'SD','TENNESSEE':'TN','TEXAS':'TX','UTAH':'UT','VERMONT':'VT',
        'VIRGINIA':'VA','WASHINGTON':'WA','WEST VIRGINIA':'WV','WISCONSIN':'WI',
        'WYOMING':'WY','DISTRICT OF COLUMBIA':'DC','DISTRICT OF COLUMBIA (DC)':'DC',
    }
    up = s.upper()
    if up in STATE_MAP: return STATE_MAP[up]
    if len(s) == 2: return s.upper()
    return s

SELF_ADMIN_VALS = {'SELF ADMINISTERED','SELF-ADMINISTERED','Self-Administered',
                   'Not Used','Not used','UNATTRIBUTED','-',''}

ADMIN_MAP = {
    'SS&C GLOBEOP':'SS&C','SS&C TECHNOLOGIES':'SS&C','SS&C ALPS':'SS&C',
    'APEX FUND SERVICES':'Apex Group','GEN II FUND SERVICES':'Gen II Fund Services',
    'STANDISH MANAGEMENT':'Standish Management','CITCO FUNDS SERVICES':'Citco Fund Services',
    'SEI FUND ADMINISTRATION':'SEI Investments','CARTA (ESHARES)':'Carta',
    'ULTIMUS LEVERPOINT FUNDS SERVICES':'Ultimus LeverPoint','STATE STREET IFS':'State Street',
    'JPMORGAN':'JP Morgan','NORTHERN TRUST':'Northern Trust',
    'NAV CONSULTING':'NAV Consulting','NAV':'NAV Consulting',
    'HEDGESERV':'HedgeServ','PETRA FUNDS GROUP':'Petra Funds Group',
    'JUNIPER SQUARE':'Juniper Square','EA RESIG LLC':'EA RESIG',
    'GP FUND SOLUTIONS':'GP Fund Solutions','ADURO ADVISORS':'Aduro Advisors',
    'ALTER DOMUS':'Alter Domus','IQ-EQ':'IQ-EQ','AZTEC GROUP':'Aztec Group',
    'TMF- LUX':'TMF Group','HC GLOBAL FUND SERVICES':'HC Global Fund Services',
    'UMB FUND SERVICES':'UMB Fund Services','MORGAN STANLEY':'Morgan Stanley',
    'CSC ENTITY SERVICES':'CSC','LANGHAM HALL':'Langham Hall',
    'E78 PARTNERS':'E78 Partners','RSM':'RSM',
    '4PINES FUND SERVICES':'4Pines Fund Services',
    'RELIANT FUND SERVICES':'Reliant Fund Services',
    'QUERCUS CANYON MANAGEMENT':'Quercus Canyon','US BANCORP':'US Bancorp',
    'TRIDENT FUND SERVICES':'Trident Trust','CLIFTON LARSON ALLEN':'CliftonLarsonAllen',
}

def norm_admin(a):
    if not a or str(a).strip() in SELF_ADMIN_VALS or str(a) in ['nan','None','']: return None
    a = str(a).strip()
    up = a.upper()
    if up in ADMIN_MAP: return ADMIN_MAP[up]
    if 'SS&C' in up: return 'SS&C'
    if up.startswith('NAV ') or up == 'NAV': return 'NAV Consulting'
    return tc(a)

# Non-sponsor firm types to exclude (only if they also have zero fund activity)
BAD_FT = {'Investment Company','Industry Association','Investment Consultant',
          'Consultant','Corporate Investor','Wealth Manager','Investment Bank',
          'Foundation','Software Company','Secondary Intermediary',
          'Private Equity Firm (Investor)','Real Estate Firm (Investor)'}

# ══════════════════════════════════════════════════════════════════
# STEP 1 — BUILD FLAGS
# ══════════════════════════════════════════════════════════════════

def build_flags():
    print("Building flags...")
    aztec_keys = set(norm(c) for c in AZTEC_CLIENTS)
    us_keys = set(norm(c) for c in AZTEC_US_CLIENTS)
    opengate_key = norm(OPENGATE_NAME)

    # CRM targets + aliases for norm mismatches
    crm_aliases = ['Halifax','Edgewaters','Windjammer Capital Investors']
    crm_keys = set(norm(t) for t in CRM_TARGETS + crm_aliases)

    crm_notes = {}
    for firm, note in CRM_NOTES.items():
        crm_notes[norm(firm)] = note

    return aztec_keys, us_keys, opengate_key, crm_keys, crm_notes

# ══════════════════════════════════════════════════════════════════
# STEP 2 — LOAD CONVERGENCE
# ══════════════════════════════════════════════════════════════════

def load_convergence():
    print("Loading Convergence manager...")
    cm = pd.read_csv(FILES['conv_managers'], low_memory=False)
    cm_us = cm[(cm['PrimaryCountry']=='UNITED STATES') &
               ((cm['Pe Pfraum Pct']>0)|(cm['Re Pfraum Pct']>0))].copy()
    cm_us['pfraum_num'] = pd.to_numeric(
        cm_us['Pfraum'].astype(str).str.replace(',',''), errors='coerce').fillna(0)
    cm_us['norm'] = cm_us['UniqueManager'].apply(norm)
    cm_lkp = {row['norm']: row for _, row in cm_us.iterrows()}
    print(f"  {len(cm_lkp):,} US managers")

    print("Computing admin tenure...")
    cf = pd.read_csv(FILES['conv_funds'], low_memory=False,
                     usecols=['UniqueManager','PrimaryCountry','MajorAssetClass',
                              'AdministratorGroup','FirstFilingDate'])
    cf_us = cf[(cf['PrimaryCountry']=='UNITED STATES') &
               (cf['MajorAssetClass'].isin(['Private Equity','Real Estate']))].copy()
    cf_us['ffp'] = pd.to_datetime(cf_us['FirstFilingDate'], format='%B %d, %Y', errors='coerce')
    cf_us.loc[cf_us['ffp']<pd.Timestamp('2000-01-01'), 'ffp'] = pd.NaT
    real_cf = cf_us[~cf_us['AdministratorGroup'].isin(['UNATTRIBUTED','SELF ADMINISTERED','-']) &
                    cf_us['AdministratorGroup'].notna()].copy()
    real_cf['an'] = real_cf['AdministratorGroup'].apply(norm_admin)
    real_cf = real_cf[real_cf['an'].notna()]

    ten_data = {}
    for mgr, g in real_cf.groupby('UniqueManager'):
        cnt = Counter(g['an'])
        cur = cnt.most_common(1)[0][0]
        dates = g[g['an']==cur]['ffp'].dropna()
        tyr = round((TODAY-dates.min()).days/365,1) if len(dates)>0 else None
        all_adm = list(set(g['an'].dropna()))
        od = g[g['an']!=cur]['ffp'].dropna()
        lsw = round((TODAY-od.max()).days/365,1) if len(od)>0 else None
        prev = None
        if len(all_adm)>1:
            others = Counter({k:v for k,v in cnt.items() if k!=cur})
            if others: prev = others.most_common(1)[0][0]
        ten_data[norm(mgr)] = {
            'cur':cur,'tyr':tyr,'all':all_adm,'sw':len(all_adm)>1,
            'lsw':lsw,'afc':int((g['an']==cur).sum()),'prev':prev
        }
    print(f"  Tenure for {len(ten_data):,} managers")
    return cm_lkp, ten_data

# ══════════════════════════════════════════════════════════════════
# STEP 3 — LOAD PREQIN
# ══════════════════════════════════════════════════════════════════

def load_preqin_funds():
    print("Loading Preqin fund-level...")
    wb = load_workbook(FILES['pq_funds'], read_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    h = list(rows[0])
    col = {name: i for i, name in enumerate(h)}
    pf_us = [r for r in rows[1:] if r[col['COUNTRY']] in ('US','Canada')]

    pf_raw = {}
    RAISING = ('Raising','First Close','Second Close','Third Close','Fourth Close')
    for r in pf_us:
        mgr = r[col['FUND MANAGER']]
        if not mgr: continue
        k = norm(str(mgr))
        if k not in pf_raw: pf_raw[k] = {'name': str(mgr), 'funds': []}
        pf_raw[k]['funds'].append(r)

    pf_agg = {}
    for k, v in pf_raw.items():
        funds = v['funds']
        statuses = [f[col['STATUS']] for f in funds if f[col['STATUS']]]
        fns = [f[col['FUND NUMBER (OVERALL)']] for f in funds if f[col['FUND NUMBER (OVERALL)']]]
        vints = [f[col['VINTAGE / INCEPTION YEAR']] for f in funds if f[col['VINTAGE / INCEPTION YEAR']]]
        aud = list(set(f[col['AUDITORS']] for f in funds
                       if f[col['AUDITORS']] and f[col['AUDITORS']] not in ('None',None)))[:3]
        laws = list(set(f[col['LAW FIRMS']] for f in funds
                        if f[col['LAW FIRMS']] and f[col['LAW FIRMS']] not in ('None',None,'Not Used')))[:3]
        pas = list(set(f[col['PLACEMENT AGENTS']] for f in funds
                       if f[col['PLACEMENT AGENTS']] and
                       f[col['PLACEMENT AGENTS']] not in ('None',None,'Not used','Not Used')))[:3]
        pf_agg[k] = {
            'cnt': len(funds),
            'raising': any(s in RAISING for s in statuses),
            'max_fn': max(fns) if fns else None,
            'lat_v': max(vints) if vints else None,
            'auditors': aud, 'laws': laws, 'pas': pas,
        }
    print(f"  Fund data for {len(pf_agg):,} managers")
    return pf_agg

def load_preqin_forecast():
    print("Loading Preqin forecast...")
    wb = load_workbook(FILES['pq_forecast'], read_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    h = list(rows[0])
    col = {name: i for i, name in enumerate(h)}
    fc_us = [r for r in rows[1:] if r[col['FUND MANAGER COUNTRY']] in ('US','Canada')]
    fc_lkp = {}
    for r in fc_us:
        k = norm(str(r[col['FUND MANAGER']] or ''))
        if k and k not in fc_lkp:
            fc_lkp[k] = {
                'nl': r[col['EST. NEXT FUND LAUNCH DATE']],
                'dp': r[col['DRY POWDER (%)']],
                'spd': r[col['DEPLOYMENT SPEED TYPE']],
                'cs': r[col['FUND CYCLE STATUS']],
                'tsl': r[col['TIME SINCE LAST LAUNCH (MONTHS)']],
                'avg': r[col['AVG TIME BETWEEN LAUNCHES (MONTHS)']],
                'fstrat': r[col['FUND STRATEGY']],
                'pa': r[col['PLACEMENT AGENTS']],
            }
    print(f"  Forecast for {len(fc_lkp):,} managers")
    return fc_lkp

def load_preqin_managers():
    print("Loading Preqin manager (full NA)...")
    wb = load_workbook(FILES['pq_managers'], read_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    h = list(rows[0])
    cp = {name: i for i, name in enumerate(h)}

    # Pre-extract column indices to avoid nested bracket parsing issues
    I = {
        'FIRM': cp['FIRM NAME'], 'COUNTRY': cp['COUNTRY'], 'FT': cp['FIRM TYPE'],
        'PE_CLOSED': cp['PE: TOTAL NO. OF FUNDS CLOSED'],
        'RE_CLOSED': cp['RE: TOTAL NO. OF FUNDS CLOSED'],
        'PE_MKT': cp['PE: TOTAL NO. OF FUNDS IN MARKET'],
        'RE_MKT': cp['RE: TOTAL NO. OF FUNDS IN MARKET'],
        'TOTAL_AUM': cp['TOTAL:ASSETS UNDER MANAGEMENT (USD MN)'],
        'PE_AUM': cp['PE: ASSETS UNDER MANAGEMENT (USD MN)'],
        'RE_AUM': cp['RE: ASSETS UNDER MANAGEMENT (USD MN)'],
        'CITY': cp['CITY'], 'STATE': cp['STATE/COUNTY'], 'YR': cp['YEAR EST.'],
        'STAFF': cp['TOTAL STAFF'],
        'PE_DP': cp['PE: ESTIMATED DRY POWDER (USD MN)'],
        'RE_DP': cp['RE: ESTIMATED DRY POWDER(USD MN)'],
        'PE_RAISED': cp['PE: TOTAL FUNDS RAISED LAST 10 YEARS (USD MN)'],
        'RE_RAISED': cp['RE: TOTAL FUNDS RAISED IN LAST 10 YEARS (USD MN)'],
        'PE_STRAT': cp['PE: STRATEGIES'], 'RE_STRAT': cp['RE: STRATEGIES'],
        'PE_GEO': cp['PE: GEOGRAPHIC EXPOSURE'], 'RE_GEO': cp['RE: GEOGRAPHIC EXPOSURE'],
        'WEB': cp['WEBSITE'], 'EMAIL': cp['EMAIL'],
        'SEC_LOCS': cp['SECONDARY LOCATIONS'],
        'PE_CO_SIZE': cp['PE: COMPANY SIZE'],
    }

    def to_f(v):
        try: return float(v) if v else 0.0
        except: return 0.0

    pm_lkp = {}
    pm_name_lkp = {}
    excluded = 0

    for r in rows[1:]:
        country = r[I['COUNTRY']]
        if country not in ('US','Canada'): continue

        ft = r[I['FT']] or ''
        pe_cl = int(r[I['PE_CLOSED']] or 0)
        re_cl = int(r[I['RE_CLOSED']] or 0)
        pe_mk = int(r[I['PE_MKT']] or 0)
        re_mk = int(r[I['RE_MKT']] or 0)
        if ft in BAD_FT and (pe_cl+re_cl+pe_mk+re_mk) == 0:
            excluded += 1
            continue

        raw_name = r[I['FIRM']]
        if not raw_name: continue
        k = norm(str(raw_name))
        if not k: continue

        pe_aum = to_f(r[I['PE_AUM']]); re_aum = to_f(r[I['RE_AUM']])
        total_aum = to_f(r[I['TOTAL_AUM']]) or pe_aum + re_aum
        city = clean(r[I['CITY']])
        state = 'Canada' if country == 'Canada' else norm_state(clean(r[I['STATE']]))
        hq = ', '.join(p for p in [city, state] if p) or None
        pe_strat = clean(r[I['PE_STRAT']]); re_strat = clean(r[I['RE_STRAT']])

        # Asset class: pe / re / both (no VC classification)
        has_re = re_aum > 0 or 'real estate' in ft.lower()
        ac = 'both' if (pe_aum > 0 and re_aum > 0) else ('re' if has_re else 'pe')

        pm_name_lkp[k] = tc(str(raw_name))
        pm_lkp[k] = {
            'total_aum_mn': total_aum, 'pe_aum_mn': pe_aum, 're_aum_mn': re_aum,
            'city': city, 'state': state, 'hq': hq, 'country': country,
            'inv_geo': clean(r[I['PE_GEO']]) or clean(r[I['RE_GEO']]),
            'yr': clean(r[I['YR']]), 'staff': clean(r[I['STAFF']]),
            'ftype': ft, 'mkt': pe_mk+re_mk, 'cls': pe_cl+re_cl,
            'dp': to_f(r[I['PE_DP']]) + to_f(r[I['RE_DP']]),
            'raised': to_f(r[I['PE_RAISED']]) + to_f(r[I['RE_RAISED']]),
            'pe_strat': pe_strat, 're_strat': re_strat,
            'web': clean(r[I['WEB']]), 'secondary_locs': clean(r[I['SEC_LOCS']]),
            'pe_company_size': clean(r[I['PE_CO_SIZE']]), 'ac_pq': ac,
        }

    print(f"  {len(pm_lkp):,} managers loaded ({excluded} non-sponsors excluded)")
    print(f"  Canadian: {sum(1 for v in pm_lkp.values() if v['country']=='Canada'):,}")
    return pm_lkp, pm_name_lkp

# ══════════════════════════════════════════════════════════════════
# STEP 4 — BUILD RECORDS
# ══════════════════════════════════════════════════════════════════

def build_records(cm_lkp, ten_data, pf_agg, fc_lkp, pm_lkp, pm_name_lkp,
                  aztec_keys, us_keys, opengate_key, crm_keys, crm_notes):
    print("Building master records...")

    all_keys = set(cm_lkp.keys()) | set(pf_agg.keys()) | set(pm_lkp.keys())
    print(f"  Universe: {len(all_keys):,}")

    CANONICAL = {}  # for manually named stubs
    for t in CRM_TARGETS:
        k = norm(t)
        CANONICAL[k] = t

    records = []
    for k in all_keys:
        cv = cm_lkp.get(k)
        pf = pf_agg.get(k) or {}
        pm = pm_lkp.get(k)
        fc = fc_lkp.get(k)
        ten = ten_data.get(k)

        # Display name
        if k in CANONICAL: name = CANONICAL[k]
        elif cv is not None: name = tc(str(cv['UniqueManager']))
        elif pm_name_lkp.get(k): name = pm_name_lkp[k]
        else: name = tc(k)
        if not name.strip(): continue

        # AUM: Preqin total first, Convergence SEC as fallback
        pm_aum = float(pm.get('total_aum_mn') or 0) if pm else 0
        conv_aum = round(float(cv['pfraum_num'])/1e6, 1) if cv is not None else 0.0
        aum = pm_aum if pm_aum > 0 else conv_aum
        aum_src = 'Preqin' if pm_aum > 0 else 'SEC-Filed RAUM (Convergence)'

        # Admin
        raw = None; is_self = False
        if cv is not None:
            pa = str(cv.get('PrimaryAdministrator','')).strip()
            if pa in SELF_ADMIN_VALS or pa == '':
                is_self = pa in ('SELF ADMINISTERED','Self-Administered')
            else: raw = pa
        if not raw and ten: raw = ten.get('cur')
        cur_adm = norm_admin(raw) if raw else None

        # Admin category
        if k in aztec_keys: cat = 'aztec'
        elif k == opengate_key: cat = 'opengate'
        elif cur_adm: cat = 'competitor'
        else: cat = 'none'

        # Contacts
        def gc(f): return clean(cv[f]) if cv is not None else None
        def ge(f):
            v = gc(f); return v if v and '@' in v else None
        cfo = tc(gc('CFO')) if gc('CFO') else None
        coo = tc(gc('COO')) if gc('COO') else None
        ceo = tc(gc('CEO')) if gc('CEO') else None
        cfe = ge('CFOEmail'); coe = ge('COOEmail'); cee = ge('CEOEmail')
        auditor_cv = tc(clean(cv['PrimaryAuditor'])) if cv is not None else None
        custodian = tc(clean(cv['PrimaryCustodian'])) if cv is not None else None
        marketer = tc(clean(cv['PrimaryMarketer'])) if cv is not None else None
        uniq_adm = int(cv['UniqueAdministrators'] or 0) if cv is not None else 0
        adm_pct = round(float(cv['PrimaryAdministrator Pfraum Pct'] or 1.0),2) if cv is not None else 1.0

        # Asset class
        if pm and pm.get('ac_pq'):
            ac = pm['ac_pq']
        else:
            pe_pct = float(cv['Pe Pfraum Pct'] or 0) if cv is not None else 0
            re_pct = float(cv['Re Pfraum Pct'] or 0) if cv is not None else 0
            if pe_pct > 20 and re_pct > 20: ac = 'both'
            elif re_pct > 20: ac = 're'
            else: ac = 'pe'

        # Geography
        country = pm.get('country','US') if pm else 'US'
        city = pm.get('city') if pm else None
        state = pm.get('state') if pm else None
        hq = pm.get('hq') if pm else None
        inv_geo = pm.get('inv_geo') if pm else None
        secondary = pm.get('secondary_locs') if pm else None

        # Year founded
        yr = None
        if pm and pm.get('yr'):
            try: yr = int(pm['yr'])
            except: pass

        # Emerging: AUM < $1B OR founded 2021+
        is_em = (0 < aum < 1000) or (yr is not None and yr >= 2021)

        # Segment
        if aum >= 50000: seg = 'large'
        elif aum >= 1000: seg = 'mid'
        elif aum > 0: seg = 'emerging'
        elif fc: seg = 'emerging'
        else: seg = 'unknown'

        # Cross-border
        cross = False
        if cv is not None:
            nu = pd.to_numeric(str(cv.get('NonUSRAUM',0)).replace(',',''), errors='coerce') or 0
            cross = nu > 0
        if not cross and inv_geo:
            cross = any(x.lower() in inv_geo.lower()
                        for x in ['Europe','Asia','Global','Middle East','Africa',
                                   'Latin America','UK','Australia','Canada'])
        if country == 'Canada': cross = True

        ten_all = ten.get('all',[]) if ten else []
        auditors_all = list(set(filter(None, [auditor_cv]+(pf.get('auditors',[]) if pf else []))))[:3]
        law_firms = (pf.get('laws',[]) if pf else [])[:3]
        pas = (pf.get('pas',[]) if pf else [])[:3]
        adm_display = cur_adm or ('Self-Administered' if is_self else None)
        all_sps = list(set(filter(None, [adm_display]+auditors_all+law_firms+pas+
                                         ([marketer] if marketer else []))))

        fc_d = fc or {}
        ten_d = ten or {}

        records.append({
            'n':name,'aum':round(aum,1),'aum_src':aum_src,'conv_aum':conv_aum,
            'ac':ac,'seg':seg,'country':country,
            'adm':adm_display,'cat':cat,'admpct':adm_pct,'uadm':uniq_adm,
            'aadm':ten_all[:5],
            'cfo':cfo,'cfe':cfe,'coo':coo,'coe':coe,'ceo':ceo,'cee':cee,
            'aud':auditors_all,'cust':custodian,'mkt':marketer,
            'law':law_firms,'pa':pas,'sps':all_sps,
            'fte':pm.get('staff') if pm else None,
            'yr':yr,'strat':pm.get('pe_strat') or pm.get('re_strat') if pm else None,
            'web':pm.get('web') if pm else None,
            'city':city,'state':state,'hq':hq,'inv_geo':inv_geo,'secondary':secondary,
            'em':is_em,'cb':cross,
            'hcv':cv is not None,'hpf':bool(pf),'hfc':fc is not None,
            'tgt':k in crm_keys,'cli':k in aztec_keys,
            'usc':k in us_keys,'og':k==opengate_key,
            'note':crm_notes.get(k,''),
            'nl':fc_d.get('nl'),'dp':fc_d.get('dp'),'spd':fc_d.get('spd'),
            'cs':fc_d.get('cs'),'tsl':fc_d.get('tsl'),'avg':fc_d.get('avg'),
            'fstrat':fc_d.get('fstrat'),'fpa':fc_d.get('pa'),
            'tyr':ten_d.get('tyr'),'afc':ten_d.get('afc',0),
            'sw':ten_d.get('sw',False),'lsw':ten_d.get('lsw'),
            'prev':ten_d.get('prev'),'tall':ten_d.get('all',[])[:5],
            'pfcnt':pf.get('cnt',0),'raising':pf.get('raising',False),
            'maxfn':pf.get('max_fn'),'latv':pf.get('lat_v'),
            'pfaud':pf.get('auditors',[]),'pflaw':pf.get('laws',[]),'pfpa':pf.get('pas',[]),
            'mktct':pm.get('mkt',0) if pm else 0,'clsct':pm.get('cls',0) if pm else 0,
            'dpmn':pm.get('dp',0) if pm else 0,'raised':pm.get('raised',0) if pm else 0,
            'pe_co_size':pm.get('pe_company_size') if pm else None,
        })

    # Sort by AUM descending
    records.sort(key=lambda x: -(x['aum'] or 0))

    # Filter to meaningful records
    meaningful = [r for r in records if r['hcv'] or r['hfc'] or r['tgt'] or r['cli']
                  or (r['aum'] or 0) >= 50 or r.get('country') == 'Canada']

    print(f"  Total: {len(records):,} | Meaningful: {len(meaningful):,}")
    print(f"  Canadian: {sum(1 for r in meaningful if r.get('country')=='Canada'):,}")
    print(f"  Targets: {sum(1 for r in meaningful if r['tgt'])}")
    print(f"  Clients: {sum(1 for r in meaningful if r['cli'])}")
    return meaningful

# ══════════════════════════════════════════════════════════════════
# STEP 5 — BUILD FILTER METADATA
# ══════════════════════════════════════════════════════════════════

def build_meta(data):
    from collections import Counter as C
    states = sorted(set(r['state'] for r in data if r.get('state')))
    cities = [c for c,_ in C(r.get('city') for r in data if r.get('city')).most_common(80)]
    admins = sorted(set(r['adm'] for r in data
                        if r.get('adm') and r['adm'] not in ('Self-Administered',None)))
    auditors = sorted(set(a for r in data
                          for a in (r.get('aud') or [])+(r.get('pfaud') or []) if a))
    laws = sorted(set(l for r in data
                      for l in (r.get('law') or [])+(r.get('pflaw') or []) if l))
    pas = sorted(set(p for r in data
                     for p in (r.get('pa') or [])+(r.get('pfpa') or [])
                     if p and p not in ('Not Used','Not used')))
    all_sps = sorted(set(sp for r in data for sp in (r.get('sps') or []) if sp))
    gv = []
    for r in data:
        if r.get('inv_geo'):
            for g in r['inv_geo'].split(','):
                g = g.strip()
                if g: gv.append(g)
    top_geos = sorted([g for g,_ in C(gv).most_common(80)])
    return {
        'states': states, 'cities': cities, 'admins': admins,
        'auditors': auditors, 'laws': laws, 'pas': pas,
        'sps': all_sps, 'geos': top_geos,
    }

# ══════════════════════════════════════════════════════════════════
# STEP 6 — INJECT INTO HTML TEMPLATE
# ══════════════════════════════════════════════════════════════════

def inject_into_html(data_str, meta, template_path=None):
    """
    Inject data into the HTML template.
    If template_path is provided, reads that file.
    Otherwise uses the embedded minimal template below.
    """
    if template_path and os.path.exists(template_path):
        print(f"Using HTML template: {template_path}")
        with open(template_path, 'r', encoding='utf-8') as f:
            html = f.read()
    else:
        print("ERROR: No HTML template found.")
        print("Place the current aztec_prospecting_v9.html in the same folder")
        print("and name it 'aztec_template.html'")
        raise FileNotFoundError("aztec_template.html not found")

    def replace_var(html, var_name, new_value):
        marker = 'const ' + var_name + '='
        idx = html.find(marker)
        if idx == -1:
            print(f"  WARNING: {var_name} not found in template")
            return html
        val_start = idx + len(marker)
        depth=0; i=val_start; in_str=False; esc_next=False
        while i < len(html):
            c = html[i]
            if esc_next: esc_next = False
            elif c == '\\' and in_str: esc_next = True
            elif c == '"' and not esc_next: in_str = not in_str
            elif not in_str:
                if c in '[{': depth += 1
                elif c in ']}':
                    depth -= 1
                    if depth == 0: i += 1; break
            i += 1
        html = html[:val_start] + new_value + html[i:]
        return html

    print("Injecting data into HTML...")
    html = replace_var(html, 'DATA', data_str)
    html = replace_var(html, 'ADMINS', json.dumps(meta['admins']))
    html = replace_var(html, 'STATES', json.dumps(meta['states']))
    html = replace_var(html, 'CITIES', json.dumps(meta['cities']))
    html = replace_var(html, 'GEOS', json.dumps(meta['geos']))
    html = replace_var(html, 'ALL_SPS', json.dumps(meta['sps']))
    return html

# ══════════════════════════════════════════════════════════════════
# MAIN
# ══════════════════════════════════════════════════════════════════

def main():
    print("=" * 60)
    print("Aztec Group — Prospecting Tool Builder")
    print("=" * 60)
    print()

    # Check input files exist
    missing = [name for name, path in FILES.items() if not os.path.exists(path)]
    if missing:
        print("ERROR: Missing input files:")
        for m in missing:
            print(f"  {FILES[m]}")
        print()
        print("Rename your Convergence/Preqin exports to match these filenames")
        print("and place them in the same folder as this script.")
        return

    # Build
    aztec_keys, us_keys, opengate_key, crm_keys, crm_notes = build_flags()
    cm_lkp, ten_data = load_convergence()
    pf_agg = load_preqin_funds()
    fc_lkp = load_preqin_forecast()
    pm_lkp, pm_name_lkp = load_preqin_managers()

    data = build_records(cm_lkp, ten_data, pf_agg, fc_lkp, pm_lkp, pm_name_lkp,
                         aztec_keys, us_keys, opengate_key, crm_keys, crm_notes)
    meta = build_meta(data)
    data_str = json.dumps(data, separators=(',',':'), default=str)

    print(f"\nData: {len(data):,} records, {len(data_str)/1024/1024:.1f} MB")

    # Find template
    template_path = os.path.join(SCRIPT_DIR, 'aztec_template.html')
    html = inject_into_html(data_str, meta, template_path)

    with open(OUTPUT_FILE, 'w', encoding='utf-8') as f:
        f.write(html)

    size_mb = len(html.encode('utf-8')) / 1024 / 1024
    print(f"\n✓ Built: {OUTPUT_FILE} ({size_mb:.1f} MB)")
    print(f"  Open in Chrome or Safari to use the tool.")
    print()

if __name__ == '__main__':
    main()
